from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from .models import  *
from django.db.models import Max
from django.forms import ModelForm
from .utils import *
from datetime import date, datetime,timedelta
from django.db import transaction
from django.utils import timezone
from django.db.models import Q
from num2words import num2words
from django.http import JsonResponse,HttpResponse
import re
from django.core.paginator import Paginator
from urllib.parse import quote
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def get_assessment_year(joining_date):
    """
    Financial Year: July 1 -> June 30
    """

    if joining_date.month >= 7:
        start_year = joining_date.year
    else:
        start_year = joining_date.year - 1

    end_year = start_year + 1

    ay, created = AssessmentYear.objects.get_or_create(
        from_date=date(start_year, 7, 1),
        to_date=date(end_year, 6, 30),
        defaults={
            "name": f"{start_year}-{end_year}"
        }
    )

    return ay


@login_required(login_url="login")
def organisation_list(request):
    organisations = Organisation.objects.all().order_by("organisation_name")
    return render(
        request,
        "membership/organisation/list.html",
        {"organisations": organisations},
    )


@login_required(login_url="login")
def organisation_create(request):

    if request.method == "POST":

        organisation = Organisation()

        organisation.organisation_name = request.POST.get("organisation_name")
        organisation.short_name = request.POST.get("short_name")

        organisation.address = request.POST.get("address")
        organisation.district = request.POST.get("district")
        organisation.state = request.POST.get("state")
        organisation.country = request.POST.get("country")

        organisation.mobile = request.POST.get("mobile")
        organisation.email = request.POST.get("email")
        organisation.website = request.POST.get("website")

        organisation.gst_no = request.POST.get("gst_no")
        organisation.pan_no = request.POST.get("pan_no")
        organisation.registration_no = request.POST.get("registration_no")

        organisation.is_active = "is_active" in request.POST

        if request.FILES.get("logo"):
            organisation.logo = request.FILES.get("logo")

        organisation.save()

        messages.success(request, "Organisation Created Successfully.")

        return redirect("organisation_list")

    return render(
        request,
        "membership/organisation/create.html",
    )

    

@login_required(login_url="login")
def organisation_edit(request, id):

    organisation = get_object_or_404(Organisation, id=id)

    if request.method == "POST":

        organisation.organisation_name = request.POST.get("organisation_name")
        organisation.short_name = request.POST.get("short_name")

        organisation.mobile = request.POST.get("mobile")
        organisation.email = request.POST.get("email")
        organisation.website = request.POST.get("website")

        organisation.country = request.POST.get("country")
        organisation.district = request.POST.get("district")
        organisation.state = request.POST.get("state")

        organisation.gst_no = request.POST.get("gst_no")
        organisation.pan_no = request.POST.get("pan_no")
        organisation.registration_no = request.POST.get("registration_no")

        organisation.address = request.POST.get("address")

        if request.FILES.get("logo"):
            organisation.logo = request.FILES.get("logo")

        organisation.save()

        messages.success(request, "Organisation Updated Successfully.")
        return redirect("organisation_list")

    return render(
        request,
        "membership/organisation/edit.html",
        {
            "organisation": organisation,
        },
    )


@login_required(login_url="login")
def organisation_delete(request, id):
    organisation = get_object_or_404(Organisation, id=id)

    if request.method == "POST":
        organisation.delete()
        messages.success(request, "Organisation deleted successfully.")
        return redirect("organisation_list")

    return render(
        request,
        "membership/organisation/delete.html",
        {"organisation": organisation},
    )

def login_view(request):

    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password,
        )

        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid Username or Password")

    return render(request, "membership/login.html")


@login_required(login_url="login")
def dashboard(request):
    return render(request, "membership/dashboard.html")


def logout_view(request):
    logout(request)
    return redirect("login")

@login_required(login_url="login")
def user_list(request):

    users = User.objects.all().order_by("username")

    return render(
        request,
        "membership/users/list.html",
        {"users": users},
    )

@login_required(login_url="login")
def user_create(request):

    if request.method == "POST":

        username = request.POST.get("username")
        first_name = request.POST.get("first_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        role = request.POST.get("role")
        password = request.POST.get("password")
        photo = request.FILES.get("photo")

        # Check duplicate username
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("user_create")

        # Create Django user
        user = User.objects.create_user(
            username=username,
            first_name=first_name,
            email=email,
            password=password,
        )

        # Create profile
        UserProfile.objects.create(
            user=user,
            role=role,
            mobile=mobile,
            photo=photo,
        )

        messages.success(request, "User created successfully.")
        return redirect("user_list")

    return render(request, "membership/users/create.html")

@login_required(login_url="login")
def user_edit(request, id):

    user = get_object_or_404(User, id=id)
    profile = get_object_or_404(UserProfile, user=user)

    if request.method == "POST":

        user.username = request.POST.get("username")
        user.first_name = request.POST.get("first_name")
        user.email = request.POST.get("email")

        profile.mobile = request.POST.get("mobile")
        profile.role = request.POST.get("role")

        if request.FILES.get("photo"):
            profile.photo = request.FILES.get("photo")

        user.save()
        profile.save()

        messages.success(request, "User updated successfully.")
        return redirect("user_list")
    
    return render(
        request,
        "membership/users/edit.html",
        {
            "user": user,
            "profile": profile,
        },
    )
@login_required(login_url="login")
def user_delete(request, id):

    user = get_object_or_404(User, id=id)

    if request.method == "POST":
        user.delete()
        messages.success(request, "User deleted successfully.")
        return redirect("user_list")

    return render(
        request,
        "membership/users/delete.html",
        {
            "user": user,
        },
    )

@login_required(login_url="login")
def members_type_list(request):

    membership_types = MembershipType.objects.all().order_by("name")

    return render(
        request,
        "membership/members_type/list.html",
        {
            "membership_types": membership_types,
        },
    )


@login_required(login_url="login")
def members_type_create(request):

    if request.method == "POST":

        membership = MembershipType()

        membership.name = request.POST.get("name")
        membership.type = request.POST.get("type")
        membership.fee = request.POST.get("fee") or 0
        membership.description = request.POST.get("description")
        membership.is_active = "is_active" in request.POST

        membership.save()

        messages.success(request, "Membership Type Created Successfully.")

        return redirect("members_type_list")

    return render(
        request,
        "membership/members_type/create.html",
    )


@login_required(login_url="login")
def members_type_edit(request, id):
    membership = get_object_or_404(MembershipType, id=id)

    if request.method == "POST":
        membership.name = request.POST.get("name")
        membership.fee = request.POST.get("fee")
        membership.validity_years = request.POST.get("validity_years")
        membership.description = request.POST.get("description")
        membership.is_active = "is_active" in request.POST

        membership.save()

        messages.success(request, "Membership Type Updated Successfully")
        return redirect("members_type_list")

    return render(
        request,
        "membership/members_type/edit.html",
        {"membership": membership},
    )


@login_required(login_url="login")
def members_type_delete(request, id):

    membership = get_object_or_404(
        MembershipType,
        id=id,
    )

    if request.method == "POST":

        membership.delete()

        messages.success(request, "Membership Type Deleted Successfully.")

        return redirect("members_type_list")

    return render(
        request,
        "membership/members_type/delete.html",
        {
            "membership": membership,
        },
    )


@login_required(login_url="login")
def member_list(request):

    members = (
        Member.objects
        .select_related(
            "organisation",
            "assessment_year",
            "membership_type",
        )
        .order_by("-member_no")
    )

    return render(
        request,
        "membership/member/list.html",
        {
            "members": members,
        },
    )

@login_required(login_url="login")
def member_create(request):
    

    membership_types = MembershipType.objects.all()
    organisations = Organisation.objects.all()

    next_member_no = ""
    next_membership_no = ""

    if organisations.exists():

        organisation = organisations.first()

        year = AssessmentYear.objects.filter(
            organisation=organisation,
            is_active=True
        ).first()

        if year:

            last = (
                Member.objects.filter(
                    organisation=organisation,
                    assessment_year=year
                )
                .order_by("-member_no")
                .first()
            )

            next_member_no = 1 if not last else last.member_no + 1
            next_membership_no = f"{year.start_year}-{next_member_no:03d}"

    if request.method == "POST":

        member = Member()
        organisation = Organisation.objects.get(
            id=request.POST.get("organisation")
        )

    

        member.organisation = organisation
        member.assessment_year = year
        membership = MembershipType.objects.get(
            id=request.POST.get("membership_type")
        )
        member.membership_type = membership

        joining_date = datetime.strptime(
            request.POST.get("joining_date"),
            "%Y-%m-%d"
        ).date()

        # April - March
        if joining_date.month >= 4:
            start_year = joining_date.year
        else:
            start_year = joining_date.year - 1

        end_year = start_year + 1

        year, created = AssessmentYear.objects.get_or_create(
            organisation=organisation,
            start_year=start_year,
            end_year=end_year,
            defaults={
                "is_active": True,
            }
        )

        member.assessment_year = year
        member.joining_date = joining_date
                        

        today = date.today()

        if membership.type == "LIFETIME":
            member.membership_valid_upto = date(2099, 12, 31)
        else:
            if today.month >= 4:
                valid_year = today.year + 1
            else:
                valid_year = today.year

            member.membership_valid_upto = date(valid_year, 3, 31)

        member.status = request.POST.get("status")

        member.owner_name = request.POST.get("owner_name")
        member.company_name = request.POST.get("company_name")
        member.brand_name = request.POST.get("brand_name")

        member.mobile = request.POST.get("mobile")
        member.whatsapp = request.POST.get("whatsapp")
        member.email = request.POST.get("email")
        member.website = request.POST.get("website")

        member.address = request.POST.get("address")
        member.district = request.POST.get("district")
        member.state = request.POST.get("state")
        member.pincode = request.POST.get("pincode")

        member.gst_no = request.POST.get("gst_no")
        member.fssai_no = request.POST.get("fssai_no")
        member.udyam_no = request.POST.get("udyam_no")
        member.factory_license_no = request.POST.get("factory_license_no")
        member.do_license_no = request.POST.get("do_license_no")
        member.trade_license_no = request.POST.get("trade_license_no")
        member.fire_noc_no = request.POST.get("fire_noc_no")
        member.pollution_no = request.POST.get("pollution_no")

        member.remarks = request.POST.get("remarks")
        

        today = date.today()

        if today.month >= 4:
            start_year = today.year
            end_year = today.year + 1
        else:
            start_year = today.year - 1
            end_year = today.year


        last_member = (
            Member.objects.filter(
                organisation=organisation,
                assessment_year=year,
            )
            .order_by("-member_no")
            .first()
        )
        
        member.member_no = 1 if not last_member else last_member.member_no + 1

        last = (
            Member.objects.filter(
                organisation=organisation,
                assessment_year=year,
            )
            .order_by("-member_no")
            .first()
        )

        if last:
            serial = int(last.membership_no.split("-")[-1]) + 1
        else:
            serial = 1

        member.membership_no = f"{year.start_year}-{serial:03d}"

      

        if request.FILES.get("photo"):
            member.photo = request.FILES.get("photo")

        if request.FILES.get("company_logo"):
            member.company_logo = request.FILES.get("company_logo")

        member.save()

        # ===============================
        # AUTO CREATE LEDGER
        # ===============================

        ledger_name = member.company_name.strip().title()

        if not Ledger.objects.filter(
            ledger_name__iexact=ledger_name
        ).exists():

            # Change this group if you create a dedicated "Members" group
            group = AccountGroup.objects.get(
                name="Current Assets"
            )

            ledger = Ledger()

            ledger.ledger_code = get_next_ledger_code(group)

            ledger.ledger_name = ledger_name

            ledger.group = group

            ledger.opening_balance = 0

            ledger.opening_type = "DEBIT"

            ledger.remarks = f"Auto Created for Member : {member.membership_no}"

            ledger.save()

        messages.success(request, "Member Created Successfully")

        return redirect("member_list")

    context = {
        "membership_types": membership_types,
        "organisations": organisations,
        "next_member_no": next_member_no,
        "next_membership_no": next_membership_no,
    }

    return render(
        request,
        "membership/member/create.html",
        context,
    )

@login_required(login_url="login")
def member_edit(request, id):

    member = get_object_or_404(Member, id=id)

    organisations = Organisation.objects.all()
    membership_types = MembershipType.objects.filter(is_active=True)

    if request.method == "POST":

        
        member.membership_type = MembershipType.objects.get(
            id=request.POST.get("membership_type")
        )

        member.joining_date = datetime.strptime(
            request.POST.get("joining_date"),
            "%Y-%m-%d"
        ).date()
        member.status = request.POST.get("status")

        member.owner_name = request.POST.get("owner_name")
        member.company_name = request.POST.get("company_name")
        member.brand_name = request.POST.get("brand_name")

        member.mobile = request.POST.get("mobile")
        member.whatsapp = request.POST.get("whatsapp")
        member.email = request.POST.get("email")
        member.website = request.POST.get("website")

        member.address = request.POST.get("address")
        member.district = request.POST.get("district")
        member.state = request.POST.get("state")
        member.pincode = request.POST.get("pincode")

        member.gst_no = request.POST.get("gst_no")
        member.fssai_no = request.POST.get("fssai_no")
        member.udyam_no = request.POST.get("udyam_no")
        member.factory_license_no = request.POST.get("factory_license_no")
        member.do_license_no = request.POST.get("do_license_no")
        member.trade_license_no = request.POST.get("trade_license_no")
        member.fire_noc_no = request.POST.get("fire_noc_no")
        member.pollution_no = request.POST.get("pollution_no")

        member.remarks = request.POST.get("remarks")

        # Membership validity
        joining_date = member.joining_date

        if member.membership_type.type == "LIFETIME":
            member.membership_valid_upto = date(2099, 12, 31)
        else:
            if joining_date.month >= 4:
                valid_year = joining_date.year + 1
            else:
                valid_year = joining_date.year

            member.membership_valid_upto = date(valid_year, 3, 31)

        if request.FILES.get("photo"):
            member.photo = request.FILES["photo"]

        if request.FILES.get("company_logo"):
            member.company_logo = request.FILES["company_logo"]

        member.save()
        

        messages.success(request, "Member Updated Successfully")

        return redirect("member_list")

    return render(
        request,
        "membership/member/edit.html",
        {
            "member": member,
            "organisations": organisations,
            "membership_types": membership_types,
        },
    )

@login_required(login_url="login")
def member_delete(request, id):
    return HttpResponse(f"Delete Member {id}")

@login_required(login_url="login")
def member_view(request, id):
    return HttpResponse(f"View Member {id}")

@login_required(login_url="login")
def fee_master_list(request):

    fees = FeeMaster.objects.all().order_by("fee_name")

    search = request.GET.get("search")

    if search:
        fees = fees.filter(
            fee_name__icontains=search
        )

    context = {
        "fees": fees,
    }

    return render(
        request,
        "membership/fee_master/list.html",
        context,
    )

@login_required(login_url="login")
def fee_master_create(request):

    organisations = Organisation.objects.all()

    if request.method == "POST":

        fee = FeeMaster()

        fee.organisation = Organisation.objects.get(
            id=request.POST.get("organisation")
        )

        fee.fee_name = request.POST.get("fee_name")
        fee.fee_type = request.POST.get("fee_type")
        fee.default_amount = request.POST.get("default_amount")
        fee.description = request.POST.get("description")

        fee.is_active = bool(
            request.POST.get("is_active")
        )

        fee.save()

        messages.success(
            request,
            "Fee Created Successfully."
        )

        return redirect("fee_master_list")

    return render(
        request,
        "membership/fee_master/create.html",
        {
            "organisations": organisations,
            "fee_types": FeeMaster.FEE_TYPES,
        },
    )


@login_required(login_url="login")
def fee_master_edit(request, id):

    fee = get_object_or_404(
        FeeMaster,
        id=id,
    )

    organisations = Organisation.objects.all()

    if request.method == "POST":

        fee.organisation = Organisation.objects.get(
            id=request.POST.get("organisation")
        )

        fee.fee_name = request.POST.get("fee_name")
        fee.fee_type = request.POST.get("fee_type")
        fee.default_amount = request.POST.get("default_amount")
        fee.description = request.POST.get("description")

        fee.is_active = bool(
            request.POST.get("is_active")
        )

        fee.save()

        messages.success(
            request,
            "Fee Updated Successfully."
        )

        return redirect("fee_master_list")

    return render(
        request,
        "membership/fee_master/edit.html",
        {
            "fee": fee,
            "organisations": organisations,
            "fee_types": FeeMaster.FEE_TYPES,
        },
    )

@login_required(login_url="login")
def fee_master_delete(request, id):

    fee = get_object_or_404(
        FeeMaster,
        id=id,
    )

    if request.method == "POST":

        fee.delete()

        messages.success(
            request,
            "Fee Deleted Successfully."
        )

        return redirect("fee_master_list")

    return render(
        request,
        "membership/fee_master/delete.html",
        {
            "fee": fee,
        },
    )

@login_required(login_url="login")
def payment_mode_list(request):

    payment_modes = PaymentMode.objects.all()

    search = request.GET.get("search")

    if search:
        payment_modes = payment_modes.filter(
            name__icontains=search
        )

    return render(
        request,
        "membership/payment_mode/list.html",
        {
            "payment_modes": payment_modes,
        },
    )

@login_required(login_url="login")
def payment_mode_create(request):

    organisations = Organisation.objects.all()

    if request.method == "POST":

        payment = PaymentMode()

        payment.organisation = Organisation.objects.get(
            id=request.POST.get("organisation")
        )

        payment.name = request.POST.get("name")
        payment.description = request.POST.get("description")

        payment.is_active = (
            request.POST.get("is_active") == "on"
        )

        payment.save()

        messages.success(
            request,
            "Payment Mode Created Successfully."
        )

        return redirect("payment_mode_list")

    return render(
        request,
        "membership/payment_mode/create.html",
        {
            "organisations": organisations,
        },
    )

@login_required(login_url="login")
def payment_mode_edit(request, id):

    payment = get_object_or_404(
        PaymentMode,
        id=id
    )

    organisations = Organisation.objects.all()

    if request.method == "POST":

        payment.organisation = Organisation.objects.get(
            id=request.POST.get("organisation")
        )

        payment.name = request.POST.get("name")
        payment.description = request.POST.get("description")

        payment.is_active = (
            request.POST.get("is_active") == "on"
        )

        payment.save()

        messages.success(
            request,
            "Payment Mode Updated Successfully."
        )

        return redirect("payment_mode_list")

    return render(
        request,
        "membership/payment_mode/edit.html",
        {
            "payment": payment,
            "organisations": organisations,
        },
    )

@login_required(login_url="login")
def payment_mode_delete(request, id):

    payment = get_object_or_404(
        PaymentMode,
        id=id
    )

    if request.method == "POST":

        payment.delete()

        messages.success(
            request,
            "Payment Mode Deleted Successfully."
        )

        return redirect("payment_mode_list")

    return render(
        request,
        "membership/payment_mode/delete.html",
        {
            "payment": payment,
        },
    )

@login_required(login_url="login")
@transaction.atomic
def receipt_create(request):

    members = Member.objects.all().order_by("company_name")

    fees = FeeMaster.objects.filter(
        is_active=True
    ).order_by("fee_name")

    payment_modes = PaymentMode.objects.filter(
        is_active=True
    ).order_by("name")

    last = Receipt.objects.order_by("-receipt_no").first()

    next_receipt_no = 1 if not last else last.receipt_no + 1

    if request.method == "POST":

        member = Member.objects.get(
            id=request.POST.get("member")
        )

        payment_mode = PaymentMode.objects.get(
            id=request.POST.get("payment_mode")
        )

        receipt = Receipt.objects.create(

            receipt_no=next_receipt_no,

            receipt_date=request.POST.get("receipt_date"),

            organisation=member.organisation,

            assessment_year=member.assessment_year,

            member=member,

            payment_mode=payment_mode,

            reference_no=request.POST.get("reference_no"),

            remarks=request.POST.get("remarks"),

            total_amount=0,
        )

        total = 0

        fee_ids = request.POST.getlist("fee_master[]")
        amounts = request.POST.getlist("amount[]")

        for fee_id, amount in zip(fee_ids, amounts):

            if not fee_id:
                continue

            fee = FeeMaster.objects.get(id=fee_id)

            ReceiptDetail.objects.create(

                receipt=receipt,

                fee_master=fee,

                amount=amount,
            )

            total += float(amount)

        receipt.total_amount = total
        receipt.save()

        messages.success(
            request,
            "Receipt Saved Successfully."
        )

        return redirect("receipt_list")

    return render(
        request,
        "membership/receipt/create.html",
        {
            "members": members,
            "fees": fees,
            "payment_modes": payment_modes,
            "next_receipt_no": next_receipt_no,
            "today": timezone.now().date(),
        },
    )

@login_required(login_url="login")
def receipt_list(request):

    receipts = (
        Receipt.objects.select_related(
            "member",
            "payment_mode",
            "organisation",
        )
        .order_by("-receipt_date", "-receipt_no")
    )

    search = request.GET.get("search")

    if search:

        receipts = receipts.filter(
            Q(receipt_no__icontains=search)
            | Q(member__owner_name__icontains=search)
            | Q(member__company_name__icontains=search)
            | Q(payment_mode__name__icontains=search)
        )

    return render(
        request,
        "membership/receipt/list.html",
        {
            "receipts": receipts,
            "search": search,
        },
    )

@login_required(login_url="login")
@transaction.atomic
def receipt_edit(request, id):

    receipt = get_object_or_404(
        Receipt.objects.prefetch_related("details"),
        id=id
    )

    members = Member.objects.order_by("company_name")

    fees = FeeMaster.objects.filter(
        is_active=True
    ).order_by("fee_name")

    payment_modes = PaymentMode.objects.filter(
        is_active=True
    ).order_by("name")

    if request.method == "POST":

        receipt.receipt_date = request.POST.get("receipt_date")

        receipt.member = Member.objects.get(
            id=request.POST.get("member")
        )

        # Update organisation & assessment year from selected member
        receipt.organisation = receipt.member.organisation
        receipt.assessment_year = receipt.member.assessment_year

        receipt.payment_mode = PaymentMode.objects.get(
            id=request.POST.get("payment_mode")
        )

        receipt.reference_no = request.POST.get("reference_no")
        receipt.remarks = request.POST.get("remarks")

        receipt.save()

        # Remove old detail rows
        receipt.details.all().delete()

        total = 0

        fee_ids = request.POST.getlist("fee_master[]")
        amounts = request.POST.getlist("amount[]")

        for fee_id, amount in zip(fee_ids, amounts):

            if not fee_id:
                continue

            fee = FeeMaster.objects.get(id=fee_id)

            ReceiptDetail.objects.create(
                receipt=receipt,
                fee_master=fee,
                amount=amount,
            )

            total += float(amount)

        receipt.total_amount = total
        receipt.save()

        messages.success(
            request,
            "Receipt Updated Successfully."
        )

        return redirect("receipt_list")

    return render(
        request,
        "membership/receipt/edit.html",
        {
            "receipt": receipt,
            "members": members,
            "fees": fees,
            "payment_modes": payment_modes,
        },
    )

@login_required(login_url="login")
def receipt_delete(request, id):

    receipt = get_object_or_404(Receipt, id=id)

    if request.method == "POST":

        receipt.delete()

        messages.success(request, "Receipt Deleted Successfully.")

        return redirect("receipt_list")

    return render(
        request,
        "membership/receipt/delete.html",
        {
            "receipt": receipt,
        },
    )
@login_required(login_url="login")
def receipt_print(request, id):

    receipt = get_object_or_404(
        Receipt.objects.select_related(
            "organisation",
            "member",
            "payment_mode",
        ).prefetch_related("details__fee_master"),
        id=id,
    )

    amount_words = (
    num2words(
        receipt.total_amount,
        lang="en_IN",
    ).title()
    + " Rupees Only"
)
    print(receipt.total_amount)
    print(amount_words)
    return render(
        request,
        "membership/receipt/print.html",
        {
            "receipt": receipt,
             "amount_words": amount_words,
        },
    )

@login_required(login_url="login")
def member_id_card(request, id):

    member = get_object_or_404(
        Member.objects.select_related(
            "organisation",
            "membership_type",
            "assessment_year",
        ),
        id=id,
    )

    context = {
        "member": member,
    }

    return render(
    request,
    "membership/member/id_card.html",
    context,
)

@login_required(login_url="login")
def member_id_card_back(request, id):

    member = get_object_or_404(
        Member.objects.select_related(
            "organisation",
            "membership_type",
            "assessment_year",
        ),
        id=id,
    )

    return render(
        request,
        "membership/member/id_card_back.html",
        {"member": member},
    )

def member_certificate(request, pk):
    member = get_object_or_404(Member, pk=pk)

    return render(
        request,
        "membership/member/certificate.html",
        {
            "member": member,
        },
    )



@login_required(login_url="login")
def member_id_card_print(request, id):

    member = get_object_or_404(
        Member.objects.select_related(
            "organisation",
            "membership_type",
            "assessment_year",
        ),
        id=id,
    )

    return render(
        request,
        "membership/member/id_card_print.html",
        {"member": member},
    )

@login_required(login_url="login")
def renewal_list(request):

    renewals = MemberRenewal.objects.select_related(
        "member"
    ).order_by("-renewal_date")


    # Search
    search = request.GET.get("search")

    if search:
        renewals = renewals.filter(
            Q(member__owner_name__icontains=search) |
            Q(renewal_no__icontains=search)
        )


    # Payment Filter
    payment_mode = request.GET.get("payment_mode")

    if payment_mode:
        renewals = renewals.filter(
            payment_mode=payment_mode
        )


    # Date Filter

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")


    if from_date:
        renewals = renewals.filter(
            renewal_date__gte=from_date
        )


    if to_date:
        renewals = renewals.filter(
            renewal_date__lte=to_date
        )


    return render(
        request,
        "membership/renewal/list.html",
        {
            "renewals": renewals,
        }
    )
@login_required(login_url="login")
def renewal_create(request):

    members = Member.objects.all().order_by("owner_name")

    if request.method == "POST":

        member = get_object_or_404(
            Member,
            id=request.POST.get("member")
        )

        assessment_year = member.assessment_year

        renewal_date = datetime.strptime(
            request.POST.get("renewal_date"),
            "%Y-%m-%d"
        ).date()

        previous_valid = member.membership_valid_upto

        if previous_valid < date.today():
            previous_valid = date.today()

        # Annual
        if member.membership_type.type == "ANNUAL":

            new_valid = date(
                previous_valid.year + 1,
                previous_valid.month,
                previous_valid.day
            )

        # Lifetime
        else:
            new_valid = previous_valid

        renewal_no = get_next_renewal_no(new_valid)

        renewal = MemberRenewal.objects.create(

            renewal_no=renewal_no,

            assessment_year=assessment_year,

            member=member,

            renewal_date=renewal_date,

            previous_valid_upto=previous_valid,

            new_valid_upto=new_valid,

            amount=request.POST.get("amount"),

            payment_mode=request.POST.get("payment_mode"),

            remarks=request.POST.get("remarks"),

            created_by=request.user,
        )

        member.membership_valid_upto = new_valid
        member.save()

        return redirect("renewal_list")
     # GET request
    assessment_year = AssessmentYear.objects.get(is_active=True)
    
    return render(
        request,
        "membership/renewal/create.html",
        {
            "members": members,
            "today": date.today(),
            "renewal_no":"Auto Generated",
            
        },
    )

def get_next_renewal_no(new_valid_upto):

    # Example:
    # new_valid_upto = 31-03-2028
    # Financial Year = 2027-28

    start_year = new_valid_upto.year - 1
    end_year = new_valid_upto.year

    fy = f"{start_year}-{str(end_year)[2:]}"

    prefix = f"REN/{fy}/"

    last = (
        MemberRenewal.objects
        .filter(renewal_no__startswith=prefix)
        .order_by("-renewal_no")
        .first()
    )

    if last:
        last_no = int(last.renewal_no.split("/")[-1])
        next_no = last_no + 1
    else:
        next_no = 1

    return f"{prefix}{next_no:04d}"


@login_required(login_url="login")
def renewal_edit(request, id):

    renewal = get_object_or_404(
        MemberRenewal.objects.select_related(
            "member",
            "member__membership_type",
            "assessment_year",
        ),
        id=id,
    )
    member = renewal.member
    membership_types = MembershipType.objects.filter(is_active=True)
    if request.method == "POST":

        renewal.renewal_date = request.POST.get("renewal_date")
        renewal.amount = request.POST.get("amount")
        renewal.payment_mode = request.POST.get("payment_mode")
        renewal.remarks = request.POST.get("remarks")

        renewal.save()

        return redirect("renewal_list")
    
    return render(
        request,
        "membership/renewal/edit.html",
        {
            "member": member,
            "renewal": renewal,
            "today": date.today(),
            "membership_types": membership_types,
        },
    )


@login_required(login_url="login")
def renewal_delete(request, id):

    renewal = get_object_or_404(MemberRenewal, id=id)

    member = renewal.member

    # Restore previous validity
    member.membership_valid_upto = renewal.previous_valid_upto
    member.save()

    renewal.delete()

    messages.success(request, "Renewal deleted successfully.")

    return redirect("renewal_list")


@login_required(login_url="login")
def renewal_print(request, id):

    renewal = get_object_or_404(
        MemberRenewal.objects.select_related(
            "member",
            "member__organisation",
            "member__membership_type",
        ),
        id=id,
    )

    amount_words = (
    num2words(renewal.amount, lang="en_IN")
    .replace(",", "")
    .title()
    + " Rupees Only"
)

    return render(
        request,
        "membership/renewal/print.html",
        {
            "renewal": renewal,
            "amount_words": amount_words,
        },
    )

@login_required(login_url="login")
def member_api(request, id):

    member = get_object_or_404(
        Member.objects.select_related("membership_type"),
        id=id,
    )

    return JsonResponse({

        "membership_no": member.membership_no,

        "owner_name": member.owner_name,

        "company_name": member.company_name,

        "membership_type": member.membership_type.name,

        "membership_kind": member.membership_type.type,

        "fee": float(member.membership_type.fee),

        "valid_upto": member.membership_valid_upto.strftime("%Y-%m-%d"),

    })
@login_required(login_url="login")
def expiry_report(request):

    today = date.today()


    total_members = Member.objects.count()


    expired_members = Member.objects.filter(
        membership_valid_upto__lt=today
    ).count()


    active_members = Member.objects.filter(
        membership_valid_upto__gte=today
    ).count()


    # Current year renewed members
    renewed_members = MemberRenewal.objects.filter(
        renewal_date__year=today.year
    ).values(
        "member"
    ).distinct().count()


    renewal_pending = active_members - renewed_members



    members = Member.objects.all().order_by(
        "membership_valid_upto"
    )


    status = request.GET.get("status")

    if status == "active":
        members = members.filter(
            membership_valid_upto__gte=today
        )

    elif status == "expired":
        members = members.filter(
            membership_valid_upto__lt=today
        )

    elif status == "renewed":
        renewed_ids = MemberRenewal.objects.filter(
            renewal_date__year=today.year
        ).values_list("member_id", flat=True)

        members = members.filter(id__in=renewed_ids)

    elif status == "pending":
        renewed_ids = MemberRenewal.objects.filter(
            renewal_date__year=today.year
        ).values_list("member_id", flat=True)

        members = members.filter(
            membership_valid_upto__gte=today
        ).exclude(
            id__in=renewed_ids
        )

     

    return render(
        request,
        "membership/report/expiry.html",
        {

            "members":members,

            "today":today,


                "total_members": total_members,
                "active_members": active_members,
                "expired_members": expired_members,
                "renewed_members": renewed_members,
                "renewal_pending": renewal_pending,


        }
    )

@login_required(login_url="login")
def whatsapp_reminder(request):

    if request.method != "POST":
        return redirect("expiry_report")

    member_ids = request.POST.getlist("members")

    if not member_ids:
        return redirect("expiry_report")
    
    # Save selected members in session
    request.session["whatsapp_members"] = member_ids
    request.session["whatsapp_index"] = 0

    members = Member.objects.filter(id__in=member_ids)

    # Preserve the user's selection order
    member_map = {str(m.id): m for m in members}
    ordered_members = [member_map[mid] for mid in member_ids if mid in member_map]

    member = ordered_members[0]

    from urllib.parse import quote

    message = (
        f"Dear {member.owner_name},\n\n"
        f"Your membership is due for renewal.\n"
        f"Please renew before "
        f"{member.membership_valid_upto.strftime('%d-%m-%Y')}.\n\n"
        f"Thank You."
    )

    whatsapp_url = (
        f"https://wa.me/91{member.mobile}"
        f"?text={quote(message)}"
    )

    return render(
        request,
        "membership/whatsapp_preview.html",
        {
            "member": member,
            "message": message,
            "whatsapp_url": whatsapp_url,
            "member_ids": ",".join(member_ids),
            "current": 1,
            "total": len(member_ids),
        },
    )


@login_required(login_url="login")
def whatsapp_preview(request):

    member_ids = request.session.get("whatsapp_members", [])
    index = request.session.get("whatsapp_index", 0)

    if not member_ids:
        return redirect("expiry_report")

    if index >= len(member_ids):
        return redirect("expiry_report")

    member = get_object_or_404(
        Member,
        id=member_ids[index]
    )

    message = (
        f"Dear {member.owner_name},\n\n"
        f"Your membership is due for renewal.\n"
        f"Please renew before "
        f"{member.membership_valid_upto.strftime('%d-%m-%Y')}.\n\n"
        f"Thank You."
    )

    whatsapp_url = (
        f"https://wa.me/91{member.mobile}"
        f"?text={quote(message)}"
    )

    return render(
        request,
        "membership/whatsapp_preview.html",
        {
            "member": member,
            "message": message,
            "whatsapp_url": whatsapp_url,
            "current": index + 1,
            "total": len(member_ids),
        },
    )

@login_required(login_url="login")
def whatsapp_next(request):

    if request.method != "POST":
        return redirect("whatsapp_preview")

    member_ids = request.session.get("whatsapp_members", [])
    index = request.session.get("whatsapp_index", 0)

    # TODO (next step): Save ReminderLog here

    index += 1
    request.session["whatsapp_index"] = index

    if index >= len(member_ids):
        request.session.pop("whatsapp_members", None)
        request.session.pop("whatsapp_index", None)

        messages.success(
            request,
            "All WhatsApp reminders completed successfully."
        )

        return redirect("expiry_report")

    return redirect("whatsapp_preview")


@login_required(login_url="login")
def account_group_list(request):

    return render(
        request,
        "membership/account/account_group_list.html"
    )


@login_required(login_url="login")
def ledger_list(request):

    ledgers = Ledger.objects.select_related(
        "group"
    ).all()

    return render(
        request,
        "membership/account/ledger_list.html",
        {
            "ledgers": ledgers
        }
    )

@login_required
def ledger_search(request):
    q = request.GET.get("q", "")

    ledgers = Ledger.objects.filter(
        ledger_name__icontains=q,
        is_active=True
    ).values(
        "id",
        "ledger_name",
        "group__name"
    )[:50]

    return JsonResponse(
        list(ledgers),
        safe=False
    )

@login_required(login_url="login")
def ledger_create(request):

    ledgers = list(
        Ledger.objects
        .select_related("group")
        .filter(is_active=True)
        .values(
            "id",
            "ledger_name",
            "group__name"
        )
    )

    groups = AccountGroup.objects.filter(
        is_active=True,
        parent__isnull=False
    ).order_by(
        "display_order",
        "name"
    )

    if request.method == "POST":

        ledger_name = request.POST.get("ledger_name").strip().title()

        if Ledger.objects.filter(
            ledger_name__iexact=ledger_name
        ).exists():

            messages.error(
                request,
                f'Ledger "{ledger_name}" already exists.'
            )

            return render(
                request,
                "membership/account/ledger_create.html",
                {
                    "groups": groups
                }
            )

        group = AccountGroup.objects.get(
            id=request.POST.get("group")
        )

        ledger_code = get_next_ledger_code(group)

        print("POST DATA =", request.POST)
        print("GROUP =", group)
        print("LEDGER CODE =", ledger_code)

        Ledger.objects.create(

            ledger_code=ledger_code,

            ledger_name=ledger_name,

            group=group,

            opening_balance=request.POST.get("opening_balance"),

            opening_type=request.POST.get("opening_type"),

            remarks=request.POST.get("remarks")

        )
        print("CREATED =", ledger.id)
        return redirect("ledger_list")

    return render(
        request,
        "membership/account/ledger_create.html",
        {
            "groups": groups
        }
    )
@login_required(login_url="login")
def ledger_edit(request,id):

    ledger=get_object_or_404(
        Ledger,
        id=id
    )

    groups=AccountGroup.objects.all()


    if request.method=="POST":

        ledger.ledger_code=request.POST.get(
            "ledger_code"
        )

        ledger.ledger_name=request.POST.get(
            "ledger_name"
        )

        ledger.group_id=request.POST.get(
            "group"
        )

        ledger.opening_balance=request.POST.get(
            "opening_balance"
        )

        ledger.opening_type=request.POST.get(
            "opening_type"
        )

        ledger.remarks=request.POST.get(
            "remarks"
        )

        ledger.save()

        return redirect(
            "ledger_list"
        )


    return render(
        request,
        "membership/account/ledger_create.html",
        {
            "ledger":ledger,
            "groups":groups
        }
    )

@login_required(login_url="login")
def ledger_delete(request,id):

    ledger=get_object_or_404(
        Ledger,
        id=id
    )

    if ledger.is_system:
        messages.error(
            request,
            "System ledger cannot be deleted"
        )

        return redirect(
            "ledger_list"
        )


    ledger.delete()

    return redirect(
        "ledger_list"
    )

def get_next_ledger_code(group):

    # Child group name
    name = group.name.lower()

    if "current assets" in name:
        start = 1000

    elif "fixed assets" in name:
        start = 1500

    elif "current liabilities" in name:
        start = 2000

    elif "membership income" in name:
        start = 4000

    elif "other income" in name:
        start = 4500

    elif "office" in name:
        start = 5000

    elif "salary" in name:
        start = 5500

    elif "electricity" in name:
        start = 6000

    else:
        start = 9000

    last = (
        Ledger.objects
        .filter(ledger_code__gte=start,
                ledger_code__lt=start + 500)
        .order_by("-ledger_code")
        .first()
    )

    if last:
        return str(int(last.ledger_code) + 1)

    return str(start + 1)

@login_required(login_url="login")
def voucher_list(request):

    vouchers = Voucher.objects.all()

    return render(
        request,
        "membership/account/voucher_list.html",
        {
            "vouchers": vouchers
        }
    )

@login_required(login_url="login")
def voucher_create(request):

    if request.method == "POST":

        print(request.POST)

        voucher = Voucher.objects.create(
        voucher_no=request.POST.get("voucher_no"),
        voucher_type=request.POST.get("voucher_type"),
        voucher_date=request.POST.get("voucher_date"),
        narration=request.POST.get("narration"),
        reference_no=request.POST.get("reference_no"),
        created_by=request.user,
    )

        ledger_ids = request.POST.getlist("ledger_id[]")
        debits = request.POST.getlist("debit[]")
        credits = request.POST.getlist("credit[]")
        remarks = request.POST.getlist("remark[]")


        for ledger_id, debit, credit, remark in zip(
            ledger_ids,
            debits,
            credits,
            remarks,
        ):
            print(ledger_id, debit, credit)

            if ledger_id:
                VoucherEntry.objects.create(
                    voucher=voucher,
                    ledger_id=ledger_id,
                    debit=debit or 0,
                    credit=credit or 0,
                    narration=remark,
                )

        messages.success(request, "Voucher Created Successfully.")
        return redirect("voucher_list")

    ledgers = list(
        Ledger.objects.select_related("group")
        .filter(is_active=True)
        .values(
            "id",
            "ledger_name",
            "group__name",
        )
        .order_by("ledger_name")
    )

    return render(
        request,
        "membership/account/voucher_create.html",
        {
            "today": date.today(),
            "voucher_no": "AUTO",
            "ledgers": ledgers,
        },
    )

@login_required
def voucher_edit(request, id):

    voucher = get_object_or_404(Voucher, id=id)

    if request.method == "POST":

        voucher.voucher_date = request.POST.get("voucher_date")
        voucher.voucher_type = request.POST.get("voucher_type")
        voucher.reference_no = request.POST.get("reference_no")
        voucher.narration = request.POST.get("narration")
        voucher.save()

        # Delete old entries
        VoucherEntry.objects.filter(voucher=voucher).delete()

        ledger_ids = request.POST.getlist("ledger_id[]")
        debits = request.POST.getlist("debit[]")
        credits = request.POST.getlist("credit[]")
        remarks = request.POST.getlist("remark[]")

        for ledger_id, debit, credit, remark in zip(
            ledger_ids,
            debits,
            credits,
            remarks,
        ):
            if ledger_id:
                VoucherEntry.objects.create(
                    voucher=voucher,
                    ledger_id=ledger_id,
                    debit=debit or 0,
                    credit=credit or 0,
                    narration=remark,
                )

        messages.success(request, "Voucher Updated Successfully.")
        return redirect("voucher_list")

    details = VoucherEntry.objects.filter(
        voucher=voucher
    ).select_related("ledger")

    context = {
        "voucher": voucher,
        "details": details,
        "ledgers": Ledger.objects.filter(is_active=True),
    }

    return render(
        request,
        "membership/account/voucher_edit.html",
        context,
    )

@login_required
def voucher_delete(request, id):

    voucher = get_object_or_404(Voucher, id=id)

    voucher.status = "CANCELLED"
    voucher.save()

    messages.success(request, "Voucher Cancelled Successfully.")
    

    return redirect("voucher_list")



@login_required(login_url="login")
def get_voucher_number(request):

    voucher_type = request.GET.get("type")

    voucher_no = get_next_voucher_no(
        voucher_type
    )

    return JsonResponse({

        "voucher_no": voucher_no

    })

@login_required
def voucher_print(request, pk):
    voucher = get_object_or_404(Voucher, pk=pk)
    organisation = Organisation.objects.first()  # adjust to however you store your single org record
    return render(request, "membership/account/voucher_print.html", {
        "voucher": voucher,
        "organisation": organisation,
    })


@login_required
def ledger_book(request, id):

    ledger = get_object_or_404(Ledger, id=id)

    entries = VoucherEntry.objects.filter(
        ledger=ledger
    ).select_related(
        "voucher",
        "ledger"
    ).order_by(
        "voucher__voucher_date",
        "voucher__voucher_no"
    )

    from_date = request.GET.get("from_date")

    to_date = request.GET.get("to_date")

    if from_date:
        entries = entries.filter(
            voucher__voucher_date__gte=from_date
        )

    if to_date:
        entries = entries.filter(
            voucher__voucher_date__lte=to_date
        )

    search = request.GET.get("search")
    export = request.GET.get("export")

    if search:
        entries = entries.filter(
            Q(voucher__voucher_no__icontains=search) |
            Q(narration__icontains=search) |
            Q(voucher__narration__icontains=search)
        )

    balance = 0

    if ledger.opening_type == "DR":
        balance = ledger.opening_balance
    else:
        balance = -ledger.opening_balance
    
    rows = []

    total_debit = 0
    total_credit = 0

    opening = ledger.opening_balance or 0

    if ledger.opening_type == "DEBIT":
        balance = opening
    else:
        balance = -opening

    total_debit = 0
    total_credit = 0

    for e in entries:

        total_debit += e.debit
        total_credit += e.credit

        balance += e.debit
        balance -= e.credit

        e.running_balance = abs(balance)
        e.balance_type = "Dr" if balance >= 0 else "Cr"

        rows.append({
            "entry": e,
            "running_balance": abs(balance),
            "balance_type": "Dr" if balance >= 0 else "Cr",
        })

    closing_balance = abs(balance)
    closing_type = "Dr" if balance >= 0 else "Cr"
    if request.GET.get("export") == "excel":
        return ledger_excel(request, id)

    if request.GET.get("export") == "pdf":
        return ledger_pdf(request, id)

    return render(
        request,
        "membership/account/ledger_book.html",
        {
            "ledger": ledger,
            "rows": rows,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "closing_balance": closing_balance,
            "closing_type": closing_type,
            "from_date": from_date,
            "to_date": to_date,
            "search": search,
            },
    )

@login_required
def ledger_print(request, id):

    ledger = get_object_or_404(Ledger, id=id)

    entries = VoucherEntry.objects.filter(
        ledger=ledger
    ).select_related(
        "voucher",
        "ledger"
    ).order_by(
        "voucher__voucher_date",
        "voucher__voucher_no"
    )

    balance = ledger.opening_balance or 0

    if ledger.opening_type == "CREDIT":
        balance = -balance

    rows = []

    total_debit = 0
    total_credit = 0

    for e in entries:

        total_debit += e.debit
        total_credit += e.credit

        balance += e.debit
        balance -= e.credit

        e.running_balance = abs(balance)
        e.balance_type = "Dr" if balance >= 0 else "Cr"

        rows.append({
            "entry": e,
        })

    return render(
        request,
        "membership/account/ledger_print.html",
        {
            "ledger": ledger,
            "rows": rows,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "closing_balance": abs(balance),
            "closing_type": "Dr" if balance >= 0 else "Cr",
        },
    )

@login_required
def ledger_pdf(request, id):

    ledger = get_object_or_404(Ledger, id=id)

    entries = (
        VoucherEntry.objects
        .filter(ledger=ledger)
        .select_related("voucher", "ledger")
        .order_by("voucher__voucher_date", "voucher__voucher_no")
    )

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        f'attachment; filename="Ledger_{ledger.ledger_code}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()

    story = []

    story.append(
        Paragraph("<b>MEMBERSHIP MANAGEMENT SYSTEM</b>", styles["Title"])
    )

    story.append(
        Paragraph(f"<b>Ledger Book : {ledger.ledger_name}</b>", styles["Heading2"])
    )

    story.append(
        Paragraph(f"Ledger Code : {ledger.ledger_code}", styles["Normal"])
    )

    story.append(Spacer(1, 8))

    data = [[
        "Date",
        "Voucher No",
        "Type",
        "Particulars",
        "Narration",
        "Debit",
        "Credit",
        "Balance",
    ]]

    opening = ledger.opening_balance or 0

    balance = opening if ledger.opening_type == "DEBIT" else -opening

    data.append([
        "",
        "",
        "",
        "Opening Balance",
        "",
        "",
        "",
        f"{ledger.opening_type} {opening:.2f}",
    ])

    total_debit = 0
    total_credit = 0

    for e in entries:

        total_debit += float(e.debit)
        total_credit += float(e.credit)

        balance += float(e.debit)
        balance -= float(e.credit)

        bal = f"{'Dr' if balance >= 0 else 'Cr'} {abs(balance):.2f}"

        data.append([
            e.voucher.voucher_date.strftime("%d-%m-%Y"),
            e.voucher.voucher_no,
            e.voucher.get_voucher_type_display(),
            e.ledger.ledger_name,
            e.narration or "",
            f"{e.debit:.2f}",
            f"{e.credit:.2f}",
            bal,
        ])

    data.append([
        "",
        "",
        "",
        "",
        "TOTAL",
        f"{total_debit:.2f}",
        f"{total_credit:.2f}",
        "",
    ])

    data.append([
        "",
        "",
        "",
        "",
        "Closing Balance",
        "",
        "",
        f"{'Dr' if balance >= 0 else 'Cr'} {abs(balance):.2f}",
    ])

    table = Table(
        data,
        colWidths=[
            22 * mm,
            28 * mm,
            28 * mm,
            55 * mm,
            60 * mm,
            25 * mm,
            25 * mm,
            30 * mm,
        ],
    )

    table.setStyle(TableStyle([

        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("FONTNAME", (0, -2), (-1, -1), "Helvetica-Bold"),

        ("BACKGROUND", (0, -2), (-1, -1), colors.beige),

        ("ALIGN", (5, 1), (-1, -1), "RIGHT"),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),

    ]))

    story.append(table)

    doc.build(story)

    return response


@login_required
def ledger_excel(request, id):

    ledger = get_object_or_404(Ledger, id=id)

    entries = (
        VoucherEntry.objects
        .filter(ledger=ledger)
        .select_related("voucher", "ledger")
        .order_by(
            "voucher__voucher_date",
            "voucher__voucher_no"
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger Book"

    # -----------------------------
    # Styles
    # -----------------------------

    title_font = Font(size=16, bold=True)

    heading_font = Font(bold=True, color="FFFFFF")

    total_font = Font(bold=True)

    blue_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    center = Alignment(horizontal="center")

    right = Alignment(horizontal="right")

    # -----------------------------
    # Heading
    # -----------------------------

    ws.merge_cells("A1:H1")
    ws["A1"] = "MEMBERSHIP MANAGEMENT SYSTEM"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Ledger Book : {ledger.ledger_name}"
    ws["A2"].font = Font(size=13, bold=True)
    ws["A2"].alignment = center

    ws["A4"] = "Ledger Code"
    ws["B4"] = ledger.ledger_code

    ws["D4"] = "Opening Balance"

    ws["E4"] = f"{ledger.opening_type} {ledger.opening_balance:.2f}"

    # -----------------------------
    # Table Header
    # -----------------------------

    headers = [

        "Date",
        "Voucher No",
        "Voucher Type",
        "Particulars",
        "Narration",
        "Debit",
        "Credit",
        "Balance",

    ]

    row = 6

    for col, text in enumerate(headers, 1):

        cell = ws.cell(row=row, column=col)

        cell.value = text

        cell.font = heading_font

        cell.fill = blue_fill

        cell.alignment = center

    # -----------------------------
    # Opening Balance
    # -----------------------------

    row += 1

    opening = ledger.opening_balance or 0

    balance = opening if ledger.opening_type == "DEBIT" else -opening

    ws.cell(row=row, column=4).value = "Opening Balance"

    ws.cell(row=row, column=8).value = (
        f"{ledger.opening_type} {opening:.2f}"
    )

    # -----------------------------
    # Transactions
    # -----------------------------

    total_debit = 0

    total_credit = 0

    for e in entries:

        row += 1

        total_debit += float(e.debit)

        total_credit += float(e.credit)

        balance += float(e.debit)

        balance -= float(e.credit)

        bal = f"{'Dr' if balance >= 0 else 'Cr'} {abs(balance):.2f}"

        ws.cell(row=row, column=1).value = e.voucher.voucher_date.strftime("%d-%m-%Y")

        ws.cell(row=row, column=2).value = e.voucher.voucher_no

        ws.cell(row=row, column=3).value = e.voucher.get_voucher_type_display()

        ws.cell(row=row, column=4).value = e.ledger.ledger_name

        ws.cell(row=row, column=5).value = e.narration

        ws.cell(row=row, column=6).value = float(e.debit)

        ws.cell(row=row, column=7).value = float(e.credit)

        ws.cell(row=row, column=8).value = bal

    # -----------------------------
    # Totals
    # -----------------------------

    row += 2

    ws.cell(row=row, column=5).value = "Total"

    ws.cell(row=row, column=6).value = total_debit

    ws.cell(row=row, column=7).value = total_credit

    ws.cell(row=row, column=5).font = total_font

    ws.cell(row=row, column=6).font = total_font

    ws.cell(row=row, column=7).font = total_font

    row += 1

    ws.cell(row=row, column=5).value = "Closing Balance"

    ws.cell(row=row, column=8).value = (
        f"{'Dr' if balance >= 0 else 'Cr'} {abs(balance):.2f}"
    )

    ws.cell(row=row, column=5).font = total_font

    ws.cell(row=row, column=8).font = total_font

    # -----------------------------
    # Auto Width
    # -----------------------------

    widths = {

        "A": 15,
        "B": 18,
        "C": 20,
        "D": 30,
        "E": 40,
        "F": 15,
        "G": 15,
        "H": 18,

    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # -----------------------------
    # Download
    # -----------------------------

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="Ledger_{ledger.ledger_code}.xlsx"'
    )

    wb.save(response)

    return response